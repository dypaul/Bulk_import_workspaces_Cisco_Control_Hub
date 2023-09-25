import requests
import openpyxl
from tqdm import tqdm

# Source of truth file
xlsx_file = "Bulk_import_workspaces.xlsx"

# Replace 'YOUR_ACCESS_TOKEN' with your authentication token
access_token = 'YOUR_ACCESS_TOKEN'

# API URL to create a location
url_create_location = 'https://webexapis.com/v1/workspaceLocations'

# API URL to create a floor 
url_create_floor = 'https://webexapis.com/v1/workspaceLocations/{locationId}/floors'

# API URL to create a workspace
url_create_workspace = 'https://webexapis.com/v1/workspaces'

# Request headers
headers = {
    'Authorization': f'Bearer {access_token}',
    'Content-Type': 'application/json'
}

# Function to get geographical coordinates (latitude and longitude) from an address
def get_lat_long(address, country_code, city_name):
    base_url = 'https://nominatim.openstreetmap.org/search'
    params = {
        'format': 'json',
        'street': address,
        'country': country_code,
        'city': city_name
    }
    response = requests.get(base_url, params=params)
    if response.status_code == 200:
        data = response.json()
        if data:
            latitude = data[0]['lat']
            longitude = data[0]['lon']
            return latitude, longitude
    return None, None

# Function to check if a location exists
def check_location_exists(display_name):
    response = requests.get(url_create_location, headers=headers)
    if response.status_code == 200:
        workspace_locations = response.json().get('items', [])
        for location in workspace_locations:
            if location.get('displayName') == display_name:
                return True, location.get('id')
    return False, None

# Function to check if a floor exists
def check_floor_exists(location_id, floor_number):
    response = requests.get(url_create_floor, headers=headers)
    if response.status_code == 200:
        floors = response.json().get('items', [])
        for floor in floors:
            if floor.get('floorNumber') == floor_number:
                return True, floor.get('id')
    return False, None

# Function to check if a workspace exists
def check_workspace_exists(workspace_name):
    response = requests.get(url_create_workspace, headers=headers)
    if response.status_code == 200:
        workspaces = response.json().get('items', [])
        for workspace in workspaces:
            if workspace.get('displayName') == workspace_name:
                return True, workspace.get('id')
    return False, None

# Function to create a location
def create_location(data):
    response = requests.post(url_create_location, headers=headers, json=data)
    if response.status_code == 200:
        return response.json().get('id')
    return None

# Function to create a floor
def create_floor(location_id, data):
    url_create_floor_for_location = url_create_floor.format(locationId=location_id)
    response = requests.post(url_create_floor_for_location, headers=headers, json=data)
    if response.status_code == 200:
        return response.json().get('id')
    return None

# Function to create a workspace
def create_workspace(data):
    response = requests.post(url_create_workspace, headers=headers, json=data)
    if response.status_code == 200 or response.status_code == 201 or response.status_code == 202:
        return True
    return False

# Read data from the Excel file
workbook = openpyxl.load_workbook(xlsx_file)
sheet = workbook.active

total_rows = sheet.max_row - 1
progress_bar = tqdm(total=total_rows, desc="Processing", unit="row", position=0)

# Iterate through the Excel file rows and create locations with floors and workspaces for each row
for row in sheet.iter_rows(min_row=2, values_only=True):
    location_name, description, address, country_code, city_name, floor_number, floor_name, workspace_name, capacity, type = row

    location_exists, location_id = check_location_exists(location_name)
    if location_exists:
        print(f"\033[0;36mThe location '{location_name}' already exists! Checking for the floor...")
        floor_exists, floor_id = check_floor_exists(location_id, floor_number)
        if floor_exists:
            print(f"\033[0;36mThe floor '{floor_number}' already exists for location '{location_name}'. Checking for the workspace...")
            workspace_exists, _ = check_workspace_exists(workspace_name)
            if workspace_exists:
                print(f"\033[0;36mThe workspace '{workspace_name}' already exists.")
            else:
                data_create_workspace = {
                    "displayName": workspace_name,
                    "locationId": location_id,
                    "floorId": floor_id,
                    "workspaceLocationId": location_id,
                    "capacity": capacity,
                    "type": type
                }
                if create_workspace(data_create_workspace):
                    print(f"\033[0;32mWorkspace '{workspace_name}' created successfully for location '{location_name}'.")
                else:
                    print(f"\033[0;31mError creating workspace for location '{location_name}'.")
        else:
            data_create_floor = {
                "floorNumber": floor_number,
                "displayName": floor_name
            }
            print(f"\033[0;32mFloor '{floor_number}' does not exist for location '{location_name}'. Creating floor...")
            floor_id = create_floor(location_id, data_create_floor)
            if floor_id is not None:
                print(f"\033[0;32mFloor '{floor_name}' created successfully for location '{location_name}'. Creating workspace...")
                data_create_workspace = {
                    "displayName": workspace_name,
                    "locationId": location_id,
                    "floorId": floor_id,
                    "workspaceLocationId": location_id,
                    "capacity": capacity,
                    "type": type
                }
                if create_workspace(data_create_workspace):
                    print(f"\033[0;32mWorkspace '{workspace_name}' created successfully for location '{location_name}'.")
                else:
                    print(f"\033[0;31mError creating workspace for location '{location_name}'.")
            else:
                print(f"\033[0;31mError creating floor for location '{location_name}'.")
    else:
        latitude, longitude = get_lat_long(address, country_code, city_name)
        if latitude is not None and longitude is not None:
            data_create_location = {
                "displayName": location_name,
                "address": address,
                "countryCode": country_code,
                "latitude": latitude,
                "longitude": longitude,
                "cityName": city_name
            }
            print(f"\033[0;37mLocation '{location_name}' does not exist. Creating location...")
            location_id = create_location(data_create_location)
            if location_id is not None:
                data_create_floor = {
                    "floorNumber": floor_number,
                    "displayName": floor_name
                }
                print(f"\033[0;37mFloor '{floor_number}' does not exist for location '{location_name}'. Creating floor...")
                floor_id = create_floor(location_id, data_create_floor)
                if floor_id is not None:
                    print(f"\033[0;37mFloor '{floor_name}' created successfully for location '{location_name}'. Creating workspace...")
                    data_create_workspace = {
                        "displayName": workspace_name,
                        "locationId": location_id,
                        "floorId": floor_id,
                        "workspaceLocationId": location_id,
                        "capacity": capacity,
                        "type": type
                    }
                    if create_workspace(data_create_workspace):
                        print(f"\033[0;32mWorkspace '{workspace_name}' created successfully for location '{location_name}'.")
                    else:
                        print(f"\033[0;31mError creating workspace for location '{location_name}'.")
                else:
                    print(f"\033[0;31mError creating floor for location '{location_name}'.")
            else:
                print(f"\033[0;31mError creating location '{location_name}'.")
        else:
            print(f"\031[0;31mUnable to obtain coordinates for location '{location_name}'. Check the address, country code, and city name.")
    # Update the progress bar
    progress_bar.update(1)
    print("\033[0;37m" + f"\r{progress_bar}", end="")
# Close the progress bar
progress_bar.close()
